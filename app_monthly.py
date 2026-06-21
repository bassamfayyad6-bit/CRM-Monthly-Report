import io
from datetime import datetime, timedelta

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="CRM Daily Report Generator", page_icon="🏭", layout="wide")

# ---------- constants ----------
HEADER_FILL = PatternFill("solid", start_color="1F3864")
COIL_CHANGE_FILL = PatternFill("solid", start_color="C6EFCE")
SHUTDOWN_FILL = PatternFill("solid", start_color="FFC7CE")
GREEN = PatternFill("solid", start_color="C6EFCE")
YELLOW = PatternFill("solid", start_color="FFEB9C")
PEACH = PatternFill("solid", start_color="FCE4D6")
RED = PatternFill("solid", start_color="FFC7CE")
THIN = Border(*[Side(style="thin")] * 4)
STOP_FILL = PatternFill("solid", start_color="F2F2F2")

# Production day runs 07:00 -> 07:00 next day.
SHIFT_START_HOUR = 7
# Stoppage classification: below = coil change, at/above = unplanned downtime (minutes)
STOP_THRESHOLD_MIN = 35


def production_day(ts):
    """The production date a timestamp belongs to (07:00 boundary).
    A coil at 03:00 on Jun-10 belongs to production day Jun-09."""
    ts = pd.Timestamp(ts)
    d = ts.normalize()
    if ts.hour < SHIFT_START_HOUR:
        d = d - pd.Timedelta(days=1)
    return d.date()


def shift_of(ts):
    """Day shift = 07:00-19:00, Night shift = 19:00-07:00."""
    h = pd.Timestamp(ts).hour
    return "Day" if SHIFT_START_HOUR <= h < SHIFT_START_HOUR + 12 else "Night"

COLS = ["#", "Coil No.", "Alloy", "Temper", "HR TH (mm)", "Target Final TH (mm)",
        "Width (mm)", "HR Weight (kg)", "Entry TH (mm)", "Required Passes",
        "Current Pass", "Exit TH after Pass (mm)", "Remaining Passes",
        "Start Time", "End Time", "Total Time (min)", "Target Speed (m/min)",
        "Actual Speed (m/min)", "Max Mill Speed (m/min)", "Speed Yield (%)",
        "Next Destination", "Notes"]

DEST_ORDER = [
    ("F.Ann", "Final annealing"),
    ("TLL", "Direct to TLL \u2014 final"),
    ("INT Trim (Slitter)", "Edge trimming"),
    ("INT Ann", "Intermediate annealing"),
    ("CM (back to CRM)", "Needs more rolling passes"),
]
NEXT_MAP = {"CM": "CM (back to CRM)", "F Ann": "F.Ann", "F ANN": "F.Ann",
            "F.Ann": "F.Ann", "INT Ann": "INT Ann", "INT ANN": "INT Ann",
            "INT Trim": "INT Trim (Slitter)", "T.L.L": "TLL", "TLL": "TLL"}


def target_speed(exit_mm):
    if exit_mm < 0.5:
        return 475
    if exit_mm <= 1:
        return 320
    if exit_mm <= 3.5:
        return 225
    return 100


def yield_fill(y):
    if y >= 95:
        return GREEN
    if y >= 85:
        return YELLOW
    if y >= 70:
        return PEACH
    return RED


def load_machine_report(file):
    df = pd.read_excel(file, sheet_name=0, header=0)
    df.columns = [str(c).strip() for c in df.columns]
    needed = {"Coil No.", "Pass No.", "Alloy Code", "Start Time", "End Time",
              "Exit Thickness [\u00b5m]", "Entry Thickness [\u00b5m]",
              "Avg. Strip Speed [m/min]", "Pup Coil", "SHIFT"}
    missing = needed - set(df.columns)
    if missing:
        raise ValueError(f"Missing columns in machine report: {missing}")
    df = df.dropna(subset=["Start Time", "End Time"])
    df["Start Time"] = pd.to_datetime(df["Start Time"])
    df["End Time"] = pd.to_datetime(df["End Time"])
    df = df.sort_values("Start Time").reset_index(drop=True)
    return df


import re as _re


def _norm_coil(x):
    """Normalize a coil id for matching against the plan.
    Strips piece markers so all pieces of a coil map to the parent:
      F2260001/3B -> F2260001/3 , 126015404/B -> 126015404 ,
      F1260017/4/A -> F1260017/4 , 1452928/B -> 1452928 , 14548520B -> 14548520"""
    s = str(x).strip().upper()
    s = _re.sub(r"/[AB]$", "", s)                 # .../B  or .../A   trailing piece
    s = _re.sub(r"([/\-]\d+)[AB]$", r"\1", s)     # .../3B -> /3
    s = _re.sub(r"(?<=\d)[AB]$", "", s)            # 126015404B -> 126015404
    return s


def _coil_variants(x):
    """All candidate keys to try when matching a machine coil to the plan."""
    s = str(x).strip().upper()
    out = {s, _norm_coil(s)}
    n = _norm_coil(s)
    # 8-digit numeric machine id -> 7-digit plan id (last digit = piece)
    if n.isdigit() and len(n) == 8:
        out.add(n[:7])
    # also try removing a trailing slash+digit piece (parent coil): 2260050/4 -> 2260050
    out.add(_re.sub(r"/\d+$", "", n))
    return {v for v in out if v}


def load_plan(file):
    plan = pd.read_excel(file, sheet_name=0, header=3)
    plan.columns = [str(c).strip().replace("\n", " ") for c in plan.columns]
    plan["_coil"] = plan["COIL Man #"].astype(str).str.strip()
    plan["_coilnorm"] = plan["_coil"].apply(_norm_coil)
    return plan


def plan_lookup(plan, coil_man, exit_mm):
    """Return dict of plan-derived fields for the matching coil row (by coil + closest targeted th)."""
    out = {"temper": "", "hr_th": "", "target_final": "", "width": "",
           "hr_weight": "", "req_passes": "", "remaining": "", "next": ""}
    if plan is None:
        return out
    try:
        cm = str(coil_man).strip()
        variants = _coil_variants(cm)
        sub = plan[(plan["_coil"].isin(variants)) | (plan["_coilnorm"].isin(variants))].copy()
        if sub.empty:
            return out
        # coil-level constants
        out["temper"] = str(sub.iloc[0].get("T.T", "")).strip()
        out["width"] = sub.iloc[0].get("Width", "")
        out["hr_weight"] = sub.iloc[0].get("Weight [Kg]", "")
        # HR TH = max raw thickness in plan for this coil
        th = pd.to_numeric(sub["TH [mm]"], errors="coerce")
        if th.notna().any():
            out["hr_th"] = round(th.max(), 2)
        # Target final TH = min targeted th for this coil
        tgt = pd.to_numeric(sub["Targeted Th."], errors="coerce")
        if tgt.notna().any():
            out["target_final"] = round(tgt.min(), 2)
        # Required passes = count of P# rows
        pass_rows = sub[sub["PASS"].astype(str).str.match(r"P\d", na=False)]
        out["req_passes"] = len(pass_rows)
        # match the row whose Targeted Th. is closest to this pass exit
        sub["_d"] = (tgt - exit_mm).abs()
        m = sub.dropna(subset=["_d"]).sort_values("_d")
        if not m.empty and m.iloc[0]["_d"] <= 0.08:
            row = m.iloc[0]
            nxt = str(row.get("NEXT", "")).strip()
            out["next"] = NEXT_MAP.get(nxt, nxt)
            # remaining = required - current pass index (best effort)
            cur = str(row.get("PASS", ""))
            if cur.startswith("P") and cur[1:].isdigit():
                out["remaining"] = max(out["req_passes"] - int(cur[1:]), 0)
        return out
    except Exception:
        return out


def _classify_stop(minutes):
    return "Coil Change Over" if minutes < STOP_THRESHOLD_MIN else "Unplanned Downtime"


def build_rows(df, plan, report_date):
    day = df[df["Start Time"].apply(production_day) == report_date].copy()
    day = day.sort_values("Start Time")

    # full production window: 07:00 of report_date -> 07:00 next day
    win_start = pd.Timestamp(report_date) + pd.Timedelta(hours=SHIFT_START_HOUR)
    win_end = win_start + pd.Timedelta(days=1)

    rows = []
    prev_end = win_start  # start counting downtime from 07:00 sharp
    for _, r in day.iterrows():
        gap = (r["Start Time"] - prev_end).total_seconds() / 60
        # only show edge/inter-coil gaps that are real (>= 1 min)
        if gap >= 1:
            rows.append({"type": "stop", "start": prev_end, "end": r["Start Time"],
                         "min": round(gap, 1), "kind": _classify_stop(gap)})
        exit_mm = round(float(r["Exit Thickness [\u00b5m]"]) / 1000, 3)
        entry_mm = round(float(r["Entry Thickness [\u00b5m]"]) / 1000, 3)
        speed = float(r["Avg. Strip Speed [m/min]"]) if pd.notna(r["Avg. Strip Speed [m/min]"]) else 0
        mill_max = r.get("Max. Mill Speed [m/min]")
        mill_max = round(float(mill_max)) if pd.notna(mill_max) else ""
        tgt = target_speed(exit_mm)
        total_min = (r["End Time"] - r["Start Time"]).total_seconds() / 60
        pl = plan_lookup(plan, r["Coil No."], exit_mm)
        rows.append({
            "type": "pass", "coil": str(r["Coil No."]),
            "shift": shift_of(r["Start Time"]),
            "alloy": str(r["Alloy Code"]),
            "pass_no": int(r["Pass No."]), "pup": r.get("Pup Coil") == 1,
            "entry": entry_mm, "exit": exit_mm,
            "start": r["Start Time"], "end": r["End Time"],
            "total": round(total_min, 1),
            "speed": round(speed), "target": tgt,
            "mill_max": mill_max,
            "yield": round(speed / tgt * 100, 1) if tgt else 0,
            "weight": pl["hr_weight"], **pl,
        })
        prev_end = r["End Time"]

    # closing gap: from last coil end to 07:00 next day
    if not day.empty:
        end_gap = (win_end - prev_end).total_seconds() / 60
        if end_gap >= 1:
            rows.append({"type": "stop", "start": prev_end, "end": win_end,
                         "min": round(end_gap, 1), "kind": _classify_stop(end_gap)})
    return rows


def build_excel(rows, report_date, engineer="", df=None, plan=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    next_day = report_date + pd.Timedelta(days=1)
    ws["A1"] = (f"CRM DAILY PRODUCTION REPORT  \u2014  {report_date.strftime('%d/%m/%Y')}"
                f"  ({report_date.strftime('%d/%m')} 07:00 \u2192 {next_day.strftime('%d/%m')} 07:00)")
    ws["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws.merge_cells("A1:V1")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = "Report Date"
    ws["C2"] = report_date.strftime("%d/%m/%Y")
    ws["E2"] = "Production Window"
    ws["G2"] = (f"{report_date.strftime('%d/%m/%Y')} 07:00  \u2192  "
                f"{next_day.strftime('%d/%m/%Y')} 07:00")
    ws["F2"] = ""
    ws["L2"] = "Shift Engineer"
    ws["N2"] = engineer
    for c in ("A2", "E2", "L2"):
        ws[c].font = Font(bold=True)

    hr = 3
    for j, c in enumerate(COLS, 1):
        cell = ws.cell(row=hr, column=j, value=c)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = THIN

    r = hr + 1
    idx = 0
    for row in rows:
        if row["type"] == "pass":
            idx += 1
            pass_lbl = f"P{row['pass_no']}" + (" (PUP)" if row["pup"] else "")
            vals = [idx, row["coil"], row["alloy"], row["temper"], row["hr_th"],
                    row["target_final"], row["width"], row["weight"], row["entry"],
                    row["req_passes"], pass_lbl, row["exit"], row["remaining"],
                    row["start"].strftime("%H:%M"), row["end"].strftime("%H:%M"),
                    row["total"], row["target"], row["speed"], row["mill_max"],
                    row["yield"], row["next"], ""]
            for j, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=j, value=v)
                cell.border = THIN
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(size=9)
            ws.cell(row=r, column=20).fill = yield_fill(row["yield"])
        else:
            is_unplanned = row["kind"] == "Unplanned Downtime"
            fill = SHUTDOWN_FILL if is_unplanned else COIL_CHANGE_FILL
            ws.cell(row=r, column=14, value=row["start"].strftime("%H:%M"))
            ws.cell(row=r, column=15, value=row["end"].strftime("%H:%M"))
            ws.cell(row=r, column=16, value=row["min"])
            ws.cell(row=r, column=21, value=row["kind"])
            for j in range(1, 23):
                cell = ws.cell(row=r, column=j)
                cell.fill = fill
                cell.border = THIN
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(size=9, bold=(j == 21),
                                 color="9C0006" if is_unplanned else "006100")
        r += 1

    last = r - 1
    ws.auto_filter.ref = f"A{hr}:W{last}"
    ws.freeze_panes = f"A{hr+1}"

    widths = [4, 13, 7, 9, 8, 10, 8, 11, 9, 9, 9, 11, 9, 9, 9, 9, 9, 9, 11, 9, 18, 14]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ---- Summary / Analysis sheet (charts + tables embedded here) ----
    s = wb.create_sheet("Summary & Analysis")
    s.sheet_view.showGridLines = False
    _next = report_date + pd.Timedelta(days=1)
    s["A1"] = (f"CRM DAILY ANALYSIS \u2014 {report_date.strftime('%d/%m/%Y')} "
               f"({report_date.strftime('%d/%m')} 07:00 \u2192 {_next.strftime('%d/%m')} 07:00)")
    s["A1"].font = Font(bold=True, size=14, color="1F3864")

    s["A3"] = "PRODUCTION SUMMARY \u2014 BY DESTINATION"
    s["A3"].font = Font(bold=True, size=12, color="1F3864")
    head_row = 4
    for j, h in enumerate(["Destination", "Coils Count", "Total HR Weight (kg)", "Note"], 1):
        c = s.cell(row=head_row, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
        c.border = THIN

    passes = [x for x in rows if x["type"] == "pass"]
    stops = [x for x in rows if x["type"] == "stop"]
    rr = head_row + 1
    total_c = total_w = 0

    def coil_wt(p):
        try:
            return float(p["weight"])
        except (ValueError, TypeError):
            return 0.0
    for dest, note in DEST_ORDER:
        ps = [p for p in passes if p["next"] == dest]
        coils = {}
        for p in ps:
            coils.setdefault(p["coil"], coil_wt(p))
        cnt = len(coils)
        wt = sum(coils.values())
        s.cell(row=rr, column=1, value=dest)
        s.cell(row=rr, column=2, value=cnt)
        s.cell(row=rr, column=3, value=round(wt))
        s.cell(row=rr, column=4, value=note)
        for j in range(1, 5):
            s.cell(row=rr, column=j).border = THIN
        total_c += cnt
        total_w += wt
        rr += 1
    s.cell(row=rr, column=1, value="TOTAL").font = Font(bold=True)
    s.cell(row=rr, column=2, value=total_c).font = Font(bold=True)
    s.cell(row=rr, column=3, value=round(total_w)).font = Font(bold=True)
    for j in range(1, 5):
        s.cell(row=rr, column=j).border = THIN

    rr += 2
    metrics_top = rr
    s.cell(row=rr, column=1, value="KEY METRICS").font = Font(bold=True, size=12, color="1F3864")
    rr += 1
    avg_act = round(sum(p["speed"] for p in passes) / len(passes)) if passes else 0
    avg_tgt = round(sum(p["target"] for p in passes) / len(passes)) if passes else 0
    avg_yld = round(sum(p["yield"] for p in passes) / len(passes), 1) if passes else 0
    day_p = [p for p in passes if p.get("shift") == "Day"]
    night_p = [p for p in passes if p.get("shift") == "Night"]
    unplanned = [st for st in stops if st.get("kind") == "Unplanned Downtime"]
    changeovers = [st for st in stops if st.get("kind") == "Coil Change Over"]
    metrics = [
        ("Total Coils Processed", len({p["coil"] for p in passes})),
        ("Total Passes", len(passes)),
        ("  \u2013 Day Shift Passes", len(day_p)),
        ("  \u2013 Night Shift Passes", len(night_p)),
        ("Total HR Weight (tons)", round(total_w / 1000, 3)),
        ("Average Actual Speed (m/min)", avg_act),
        ("Average Target Speed (m/min)", avg_tgt),
        ("Overall Speed Yield (%)", avg_yld),
        ("Change-overs (< 20 min)", f"{len(changeovers)}  ({sum(st['min'] for st in changeovers):.0f} min)"),
        ("Unplanned Downtime (\u2265 20 min)", f"{len(unplanned)}  ({sum(st['min'] for st in unplanned):.0f} min)"),
        ("Total Stoppage (min)", round(sum(st["min"] for st in stops), 1)),
    ]
    for k, v in metrics:
        s.cell(row=rr, column=1, value=k).font = Font(bold=not k.startswith("  "))
        s.cell(row=rr, column=3, value=v)
        rr += 1

    s.column_dimensions["A"].width = 30
    s.column_dimensions["B"].width = 14
    s.column_dimensions["C"].width = 20
    s.column_dimensions["D"].width = 26

    # ---- Embedded charts + speed-group tables (need df) ----
    if df is not None:
        try:
            _embed_analysis(s, wb, df, plan, report_date, anchor_row=rr + 2)
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


SPEED_GROUPS = [
    ("180\u2013350 \u00b5m", 180, 350, 500),
    ("350\u2013500 \u00b5m", 350, 500, 500),
    ("500\u2013800 \u00b5m", 500, 800, 450),
    ("800\u20131800 \u00b5m", 800, 1800, 350),
    ("1800\u20135000 \u00b5m", 1800, 5000, 100),
]
HT_TARGET = 8
NAVY = "#1F3864"


def _prep(df):
    d = df.copy()
    d["ShiftName"] = d["Start Time"].apply(shift_of)
    d["Date"] = d["Start Time"].apply(production_day)
    d["ExitUm"] = pd.to_numeric(d["Exit Thickness [\u00b5m]"], errors="coerce")
    d["Len"] = pd.to_numeric(d["Coil Length [m]"], errors="coerce").fillna(0)
    d["Roll"] = pd.to_numeric(d["Rolling time"], errors="coerce").fillna(0)
    ht_raw = pd.to_numeric(d["Handling Time [min]"], errors="coerce")
    d["HT"] = ht_raw.where(ht_raw <= 20)
    d["UnplannedStop"] = ht_raw.where(ht_raw > 20)
    d["Wt"] = pd.to_numeric(d["Measured Weight [t]"], errors="coerce").fillna(0)
    d["Pup"] = pd.to_numeric(d["Pup Coil"], errors="coerce").fillna(0).astype(int)

    def grp(u):
        for lbl, lo, hi, tgt in SPEED_GROUPS:
            if lo <= u < hi or (hi == 5000 and u == 5000):
                return lbl
        return None

    def tgt(u):
        for lbl, lo, hi, t in SPEED_GROUPS:
            if lo <= u < hi or (hi == 5000 and u == 5000):
                return t
        return None

    d["Grp"] = d["ExitUm"].apply(lambda u: grp(u) if pd.notna(u) else None)
    d["TgtSpd"] = d["ExitUm"].apply(lambda u: tgt(u) if pd.notna(u) else None)
    d["TheoTime"] = d.apply(lambda r: r["Len"] / r["TgtSpd"] if r["TgtSpd"] else None, axis=1)
    d["RunPerf"] = d.apply(
        lambda r: 100 * r["TheoTime"] / r["Roll"] if r["TheoTime"] and r["Roll"] > 0 else None, axis=1)
    return d


def _table(ax, col_labels, rows, title=None, col_widths=None, bold_last=False, red_col=None):
    ax.axis("off")
    if title:
        ax.set_title(title, fontsize=10, fontweight="bold", loc="left", color=NAVY)
    t = ax.table(cellText=rows, colLabels=col_labels, loc="upper left", cellLoc="left")
    has_header = col_labels is not None
    t.auto_set_font_size(False)
    t.set_fontsize(7.5)
    t.scale(1, 1.35)
    for (r, c), cell in t.get_celld().items():
        cell.set_edgecolor("#CCCCCC")
        if r == 0 and has_header:
            cell.set_facecolor("#2F4C6E")
            cell.set_text_props(color="white", fontweight="bold")
        elif not has_header and c % 2 == 0:
            cell.set_text_props(fontweight="bold")
        if bold_last and has_header and r == len(rows):
            cell.set_facecolor("#EFEFEF")
            cell.set_text_props(fontweight="bold")
        if red_col is not None and r > 0 and c == red_col:
            cell.set_text_props(color="#C00000")
    if col_widths:
        for c, w in enumerate(col_widths):
            for r in range(len(rows) + 1):
                t[r, c].set_width(w)


def _group_table_rows(sub):
    rows = []
    tot_theo = tot_roll = tot_coils = 0
    spd_w = []
    for lbl, lo, hi, tgt in SPEED_GROUPS:
        g = sub[sub["Grp"] == lbl]
        g = g[(g["Roll"] > 0) & g["TheoTime"].notna()]
        if g.empty:
            rows.append([lbl, tgt, "-", "0.0", "0.0", "+0.0", "-", 0])
            continue
        theo, roll = g["TheoTime"].sum(), g["Roll"].sum()
        spd = (g["Len"].sum() / roll) if roll else 0
        rows.append([lbl, tgt, round(spd), f"{theo:.1f}", f"{roll:.1f}",
                     f"+{roll - theo:.1f}", f"{100 * theo / roll:.1f}%", len(g)])
        tot_theo += theo; tot_roll += roll; tot_coils += len(g); spd_w.append((spd, roll))
    avg_spd = round(sum(s * w for s, w in spd_w) / sum(w for _, w in spd_w)) if spd_w else 0
    rows.append(["Total", "\u2014", avg_spd, f"{tot_theo:.1f}", f"{tot_roll:.1f}",
                 f"+{tot_roll - tot_theo:.1f}",
                 f"{100 * tot_theo / tot_roll:.1f}%" if tot_roll else "-", tot_coils])
    perf = 100 * tot_theo / tot_roll if tot_roll else 0
    return rows, perf


def _fig_to_img(fig):
    """Render a matplotlib figure to an openpyxl image."""
    from openpyxl.drawing.image import Image as XLImage
    b = io.BytesIO()
    fig.savefig(b, format="png", dpi=130, bbox_inches="tight")
    plt.close(fig)
    b.seek(0)
    return XLImage(b)


def _embed_analysis(s, wb, df, plan, report_date, anchor_row):
    """Draw per-coil charts + speed-group tables directly into the Summary sheet."""
    d = _prep(df)
    if report_date is not None:
        d = d[d["Date"] == report_date].copy()
    if d.empty:
        return
    d = d.sort_values("Start Time").reset_index(drop=True)

    row = anchor_row

    # ---- Chart 1: Run Performance per coil ----
    s.cell(row=row, column=1,
           value="RUN PERFORMANCE (%) PER COIL").font = Font(bold=True, size=12, color="1F3864")
    row += 1
    rc = d[(d["Roll"] > 0) & d["RunPerf"].notna()].reset_index(drop=True)
    if len(rc):
        fig, ax = plt.subplots(figsize=(11, 3.4))
        colors = ["#1F77B4" if sh == "Day" else "#FF7F0E" for sh in rc["ShiftName"]]
        ax.bar(range(len(rc)), rc["RunPerf"], color=colors, width=0.8)
        ax.axhline(100, ls="--", c="#2E7D32", lw=1.4)
        ax.set_ylabel("Run Perf (%)")
        ax.set_xticks(range(len(rc)))
        ax.set_xticklabels([str(c)[-5:] for c in rc["Coil No."]], rotation=90, fontsize=5)
        ax.set_xlim(-0.6, len(rc) - 0.4)
        from matplotlib.patches import Patch
        ax.legend(handles=[Patch(color="#1F77B4", label="Day"),
                           Patch(color="#FF7F0E", label="Night"),
                           plt.Line2D([0], [0], ls="--", c="#2E7D32", label="Target 100%")],
                  loc="upper right", fontsize=7, ncol=3)
        s.add_image(_fig_to_img(fig), f"A{row}")
    row += 18

    # ---- Chart 2: Change-over time per coil (target 8 min) ----
    s.cell(row=row, column=1,
           value="CHANGE-OVER TIME PER COIL (target 8 min)").font = Font(bold=True, size=12, color="1F3864")
    row += 1
    hc = d[d["HT"].notna()].reset_index(drop=True)
    if len(hc):
        fig, ax = plt.subplots(figsize=(11, 3.4))
        colors = ["#1F77B4" if sh == "Day" else "#FF7F0E" for sh in hc["ShiftName"]]
        bars = ax.bar(range(len(hc)), hc["HT"], color=colors, width=0.8)
        ax.axhline(HT_TARGET, ls="--", c="#C00000", lw=1.6)
        for b, v in zip(bars, hc["HT"]):
            if v > HT_TARGET:
                b.set_edgecolor("#C00000"); b.set_linewidth(0.8)
        ax.set_ylabel("Change-over (min)")
        ax.set_xticks(range(len(hc)))
        ax.set_xticklabels([str(c)[-5:] for c in hc["Coil No."]], rotation=90, fontsize=5)
        ax.set_xlim(-0.6, len(hc) - 0.4)
        from matplotlib.patches import Patch
        ax.legend(handles=[Patch(color="#1F77B4", label="Day"),
                           Patch(color="#FF7F0E", label="Night"),
                           plt.Line2D([0], [0], ls="--", c="#C00000", label=f"Target {HT_TARGET} min")],
                  loc="upper right", fontsize=7, ncol=3)
        s.add_image(_fig_to_img(fig), f"A{row}")
    row += 18

    # ---- Speed-group tables: Day / Night / Total ----
    cols = ["Thickness Range", "Target Speed", "Actual Avg Speed", "Theoretical Time (min)",
            "Rolling Time (min)", "Loss (min)", "Run Perf (%)", "Coils"]
    for label, sub in [("DAY SHIFT", d[d["ShiftName"] == "Day"]),
                       ("NIGHT SHIFT", d[d["ShiftName"] == "Night"]),
                       ("TOTAL", d)]:
        grp_rows, perf = _group_table_rows(sub)
        s.cell(row=row, column=1,
               value=f"RUN PERFORMANCE BY SPEED GROUP — {label} (Run Perf: {perf:.1f}%)"
               ).font = Font(bold=True, size=11, color="1F3864")
        row += 1
        for j, h in enumerate(cols, 1):
            c = s.cell(row=row, column=j, value=h)
            c.fill = HEADER_FILL
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border = THIN
        row += 1
        for gr in grp_rows:
            is_total = gr[0] == "Total"
            for j, v in enumerate(gr, 1):
                c = s.cell(row=row, column=j, value=v)
                c.border = THIN
                c.alignment = Alignment(horizontal="center")
                c.font = Font(size=9, bold=is_total)
                if is_total:
                    c.fill = PatternFill("solid", start_color="EFEFEF")
                if j == 6:  # Loss column
                    c.font = Font(size=9, bold=is_total, color="C00000")
            row += 1
        row += 2



# ============================================================
#   MONTHLY AGGREGATION  (reuses daily core above)
# ============================================================
import calendar


def month_label(y, m):
    return f"{calendar.month_name[m]} {y}"


def build_monthly_excel(df, plan, year, month, engineer=""):
    """Build a full monthly report: a detailed sheet (all coils, day by day,
    same columns as the daily report) + a Summary & Analysis sheet aggregated
    over the whole month."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    pdays = df["Start Time"].apply(production_day)
    month_days = sorted({d for d in pdays if d.year == year and d.month == month})
    if not month_days:
        raise ValueError("No production days for the selected month.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Report"

    ws["A1"] = f"CRM MONTHLY PRODUCTION REPORT  \u2014  {month_label(year, month)}"
    ws["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws.merge_cells("A1:U1")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = "Month"
    ws["C2"] = month_label(year, month)
    ws["E2"] = "Production Days"
    ws["G2"] = len(month_days)
    ws["L2"] = "Prepared By"
    ws["N2"] = engineer
    for c in ("A2", "E2", "L2"):
        ws[c].font = Font(bold=True)

    hr = 3
    # Detailed columns = daily columns + a leading Date column
    month_cols = ["Date", "Shift"] + COLS[1:]  # replace '#' with Date, add Shift
    for j, c in enumerate(month_cols, 1):
        cell = ws.cell(row=hr, column=j, value=c)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = THIN

    r = hr + 1
    # collect aggregated structures
    all_rows = []          # all daily rows combined (for monthly analysis)
    daily_summaries = []   # per-day metric dict
    for d in month_days:
        rows = build_rows(df, plan, d)
        all_rows.extend([(d, x) for x in rows])
        passes = [x for x in rows if x["type"] == "pass"]
        stops = [x for x in rows if x["type"] == "stop"]
        for row in rows:
            if row["type"] == "pass":
                pass_lbl = f"P{row['pass_no']}" + (" (PUP)" if row["pup"] else "")
                vals = [d.strftime("%d/%m"), row.get("shift", ""), row["coil"], row["alloy"],
                        row["temper"], row["hr_th"], row["target_final"], row["width"],
                        row["weight"], row["entry"], row["req_passes"], pass_lbl, row["exit"],
                        row["remaining"], row["start"].strftime("%H:%M"),
                        row["end"].strftime("%H:%M"), row["total"], row["target"],
                        row["speed"], row["mill_max"], row["yield"], row["next"], ""]
                for j, v in enumerate(vals, 1):
                    cell = ws.cell(row=r, column=j, value=v)
                    cell.border = THIN
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(size=9)
                ws.cell(row=r, column=21).fill = yield_fill(row["yield"])
            else:
                is_unpl = row["kind"] == "Unplanned Downtime"
                fill = SHUTDOWN_FILL if is_unpl else COIL_CHANGE_FILL
                ws.cell(row=r, column=1, value=d.strftime("%d/%m"))
                ws.cell(row=r, column=15, value=row["start"].strftime("%H:%M"))
                ws.cell(row=r, column=16, value=row["end"].strftime("%H:%M"))
                ws.cell(row=r, column=17, value=row["min"])
                ws.cell(row=r, column=22, value=row["kind"])
                for j in range(1, 24):
                    cell = ws.cell(row=r, column=j)
                    cell.fill = fill
                    cell.border = THIN
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(size=9, bold=(j == 22),
                                     color="9C0006" if is_unpl else "006100")
            r += 1

        # per-day summary metrics
        tw = 0.0
        coils = set()
        for p in passes:
            coils.add(p["coil"])
            try:
                tw += float(p["weight"])
            except (ValueError, TypeError):
                pass
        # weight once per coil
        cw = {}
        for p in passes:
            try:
                cw.setdefault(p["coil"], float(p["weight"]))
            except (ValueError, TypeError):
                cw.setdefault(p["coil"], 0.0)
        daily_summaries.append({
            "date": d,
            "coils": len(coils),
            "passes": len(passes),
            "weight_t": sum(cw.values()) / 1000,
            "avg_yield": round(sum(p["yield"] for p in passes) / len(passes), 1) if passes else 0,
            "changeover": sum(1 for s in stops if s["kind"] == "Coil Change Over"),
            "unplanned": sum(1 for s in stops if s["kind"] == "Unplanned Downtime"),
            "downtime_min": round(sum(s["min"] for s in stops if s["kind"] == "Unplanned Downtime")),
        })

    last = r - 1
    ws.auto_filter.ref = f"A{hr}:W{last}"
    ws.freeze_panes = f"A{hr+1}"
    widths = [8, 7, 13, 7, 9, 8, 10, 8, 11, 9, 9, 11, 9, 9, 9, 9, 9, 9, 11, 9, 9, 18, 12]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ---- Monthly Summary & Analysis sheet ----
    _build_monthly_summary(wb, df, plan, year, month, month_days,
                           all_rows, daily_summaries)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_monthly_summary(wb, df, plan, year, month, month_days, all_rows, daily_summaries):
    from openpyxl.styles import Alignment, Font, PatternFill
    s = wb.create_sheet("Summary & Analysis")
    s.sheet_view.showGridLines = False
    s["A1"] = f"CRM MONTHLY ANALYSIS \u2014 {month_label(year, month)}"
    s["A1"].font = Font(bold=True, size=14, color="1F3864")

    passes = [x for (_, x) in all_rows if x["type"] == "pass"]
    stops = [x for (_, x) in all_rows if x["type"] == "stop"]

    # KEY METRICS
    s["A3"] = "MONTHLY KEY METRICS"
    s["A3"].font = Font(bold=True, size=12, color="1F3864")
    coil_w = {}
    for p in passes:
        try:
            coil_w.setdefault(p["coil"], float(p["weight"]))
        except (ValueError, TypeError):
            coil_w.setdefault(p["coil"], 0.0)
    day_p = [p for p in passes if p.get("shift") == "Day"]
    night_p = [p for p in passes if p.get("shift") == "Night"]
    co = [s_ for s_ in stops if s_["kind"] == "Coil Change Over"]
    un = [s_ for s_ in stops if s_["kind"] == "Unplanned Downtime"]
    metrics = [
        ("Production Days", len(month_days)),
        ("Total Coils Processed", len({p["coil"] for p in passes})),
        ("Total Passes", len(passes)),
        ("  \u2013 Day Shift Passes", len(day_p)),
        ("  \u2013 Night Shift Passes", len(night_p)),
        ("Total HR Weight (tons)", round(sum(coil_w.values()) / 1000, 1)),
        ("Avg Speed Yield (%)", round(sum(p["yield"] for p in passes) / len(passes), 1) if passes else 0),
        ("Change-overs (< 20 min)", f"{len(co)}  ({sum(x['min'] for x in co):.0f} min)"),
        ("Unplanned Downtime (\u2265 20 min)", f"{len(un)}  ({sum(x['min'] for x in un):.0f} min)"),
        ("Total Downtime (hours)", round(sum(x["min"] for x in un) / 60, 1)),
    ]
    rr = 4
    for k, v in metrics:
        s.cell(row=rr, column=1, value=k).font = Font(bold=not k.startswith("  "))
        s.cell(row=rr, column=3, value=v)
        rr += 1

    # BY DESTINATION (month total)
    rr += 1
    s.cell(row=rr, column=1, value="MONTHLY PRODUCTION BY DESTINATION").font = Font(bold=True, size=12, color="1F3864")
    rr += 1
    for j, h in enumerate(["Destination", "Coils Count", "Total HR Weight (t)", "Note"], 1):
        c = s.cell(row=rr, column=j, value=h)
        c.fill = HEADER_FILL; c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center"); c.border = THIN
    rr += 1
    for dest, note in DEST_ORDER:
        ps = [p for p in passes if p["next"] == dest]
        coils = {}
        for p in ps:
            try:
                coils.setdefault(p["coil"], float(p["weight"]))
            except (ValueError, TypeError):
                coils.setdefault(p["coil"], 0.0)
        s.cell(row=rr, column=1, value=dest)
        s.cell(row=rr, column=2, value=len(coils))
        s.cell(row=rr, column=3, value=round(sum(coils.values()) / 1000, 1))
        s.cell(row=rr, column=4, value=note)
        for j in range(1, 5):
            s.cell(row=rr, column=j).border = THIN
        rr += 1

    s.column_dimensions["A"].width = 32
    s.column_dimensions["B"].width = 14
    s.column_dimensions["C"].width = 20
    s.column_dimensions["D"].width = 26

    # ---- Charts ----
    chart_row = rr + 2
    ds = daily_summaries
    dates = [d["date"].strftime("%d") for d in ds]

    # Chart 1: daily coils + weight
    fig, ax = plt.subplots(figsize=(11, 3.2))
    ax.bar(range(len(ds)), [d["coils"] for d in ds], color="#1F3864")
    ax.set_ylabel("Coils / day"); ax.set_title("Daily Coil Count", fontsize=11,
                  fontweight="bold", loc="left", color="#1F3864")
    ax.set_xticks(range(len(ds))); ax.set_xticklabels(dates, fontsize=7)
    s.cell(row=chart_row, column=1, value="DAILY PRODUCTION TREND").font = Font(bold=True, size=12, color="1F3864")
    s.add_image(_fig_to_img(fig), f"A{chart_row + 1}")
    chart_row += 19

    # Chart 2: daily speed yield %
    fig, ax = plt.subplots(figsize=(11, 3.2))
    yl = [d["avg_yield"] for d in ds]
    colors = ["#2E7D32" if y >= 95 else "#F9A825" if y >= 85 else "#EF6C00" if y >= 70 else "#C62828" for y in yl]
    ax.bar(range(len(ds)), yl, color=colors)
    ax.axhline(95, ls="--", c="#2E7D32", lw=1.2)
    ax.set_ylabel("Avg Speed Yield (%)"); ax.set_title("Daily Average Speed Yield", fontsize=11,
                  fontweight="bold", loc="left", color="#1F3864")
    ax.set_xticks(range(len(ds))); ax.set_xticklabels(dates, fontsize=7)
    s.cell(row=chart_row, column=1, value="DAILY SPEED YIELD").font = Font(bold=True, size=12, color="1F3864")
    s.add_image(_fig_to_img(fig), f"A{chart_row + 1}")
    chart_row += 19

    # Chart 3: daily downtime (min)
    fig, ax = plt.subplots(figsize=(11, 3.2))
    ax.bar(range(len(ds)), [d["downtime_min"] for d in ds], color="#C62828")
    ax.set_ylabel("Unplanned Downtime (min)"); ax.set_title("Daily Unplanned Downtime", fontsize=11,
                  fontweight="bold", loc="left", color="#1F3864")
    ax.set_xticks(range(len(ds))); ax.set_xticklabels(dates, fontsize=7)
    s.cell(row=chart_row, column=1, value="DAILY DOWNTIME").font = Font(bold=True, size=12, color="1F3864")
    s.add_image(_fig_to_img(fig), f"A{chart_row + 1}")
    chart_row += 19

    # Speed-group tables (month total, day, night)
    d_all = _prep(df)
    d_all = d_all[d_all["Date"].apply(lambda x: x.year == year and x.month == month)]
    cols = ["Thickness Range", "Target Speed", "Actual Avg Speed", "Theoretical Time (min)",
            "Rolling Time (min)", "Loss (min)", "Run Perf (%)", "Coils"]
    for label, sub in [("DAY SHIFT", d_all[d_all["ShiftName"] == "Day"]),
                       ("NIGHT SHIFT", d_all[d_all["ShiftName"] == "Night"]),
                       ("MONTH TOTAL", d_all)]:
        grp_rows, perf = _group_table_rows(sub)
        s.cell(row=chart_row, column=1,
               value=f"RUN PERFORMANCE BY SPEED GROUP \u2014 {label} (Run Perf: {perf:.1f}%)"
               ).font = Font(bold=True, size=11, color="1F3864")
        chart_row += 1
        for j, h in enumerate(cols, 1):
            c = s.cell(row=chart_row, column=j, value=h)
            c.fill = HEADER_FILL; c.font = Font(bold=True, color="FFFFFF", size=9)
            c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = THIN
        chart_row += 1
        for gr in grp_rows:
            is_total = gr[0] == "Total"
            for j, v in enumerate(gr, 1):
                c = s.cell(row=chart_row, column=j, value=v)
                c.border = THIN; c.alignment = Alignment(horizontal="center")
                c.font = Font(size=9, bold=is_total, color="C00000" if j == 6 else "000000")
                if is_total:
                    c.fill = PatternFill("solid", start_color="EFEFEF")
            chart_row += 1
        chart_row += 2

# ---------- UI ----------
st.set_page_config(page_title="CRM Monthly Report", page_icon="📅", layout="wide")
st.title("📅 CRM Monthly Report Generator")
st.caption("Upload one or more L2 machine reports, pick a month, and get a full monthly report + analysis.")


@st.cache_data(show_spinner=False)
def _read_machine_m(data: bytes, name: str):
    b = io.BytesIO(data); b.name = name
    return load_machine_report(b)


@st.cache_data(show_spinner=False)
def _read_plan_m(data: bytes, name: str):
    b = io.BytesIO(data); b.name = name
    return load_plan(b)


st.subheader("1) Upload L2 machine report(s)")
st.caption("You can add several files (e.g. one per period). They will be merged automatically.")
machine_ups = st.file_uploader("L2 Machine Reports", type=["xls", "xlsx"],
                               accept_multiple_files=True, key="m_machine")
plan_up = st.file_uploader("Rolling Plan (optional, for Destination & plan columns)",
                           type=["xlsx"], key="m_plan")

# persist
if machine_ups:
    st.session_state["m_files"] = [(f.name, f.getvalue()) for f in machine_ups]
if plan_up is not None:
    st.session_state["m_plan_bytes"] = plan_up.getvalue()
    st.session_state["m_plan_name"] = plan_up.name

mfiles = st.session_state.get("m_files", [])
pbytes = st.session_state.get("m_plan_bytes")

if mfiles:
    st.success(f"✓ {len(mfiles)} machine file(s) loaded: " + ", ".join(n for n, _ in mfiles))
if pbytes:
    st.success(f"✓ Rolling plan loaded: {st.session_state.get('m_plan_name','')}")

if mfiles:
    try:
        # merge all machine files
        frames = [_read_machine_m(data, name) for name, data in mfiles]
        df = pd.concat(frames, ignore_index=True).drop_duplicates()
        df = df.sort_values("Start Time").reset_index(drop=True)
        plan = _read_plan_m(pbytes, st.session_state.get("m_plan_name", "p.xlsx")) if pbytes else None
        if plan is None:
            st.warning("⚠️ No rolling plan uploaded — Temper / HR TH / Next Destination will be blank.")

        # available months
        pdays = df["Start Time"].apply(production_day)
        months = sorted({(d.year, d.month) for d in pdays}, reverse=True)

        st.subheader("2) Pick the month")
        sel = st.selectbox("Month", months, format_func=lambda ym: month_label(*ym), key="m_month")
        engineer = st.text_input("Prepared By (optional)", key="m_eng")

        ndays = len({d for d in pdays if (d.year, d.month) == sel})
        ncoils = sum(1 for d in pdays if (d.year, d.month) == sel)
        st.info(f"📊 {month_label(*sel)} — {ncoils} coils across {ndays} production days.")

        if st.button("✅ OK — Generate Monthly Report", type="primary"):
            with st.spinner("Building monthly report… this can take a moment for a full month."):
                xb = build_monthly_excel(df, plan, sel[0], sel[1], engineer)
            st.session_state["m_result"] = {"xlsx": xb.getvalue(), "month": sel,
                                            "days": ndays, "coils": ncoils}

        res = st.session_state.get("m_result")
        if res and res["month"] == sel:
            st.success(f"Done — {month_label(*sel)}: {res['coils']} coils, {res['days']} days.")
            st.download_button("⬇️ Download Monthly Report (Excel)", res["xlsx"],
                               file_name=f"CRM_Monthly_Report_{sel[0]}_{sel[1]:02d}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        st.error(f"Error: {e}")
        st.exception(e)
else:
    st.info("⬆️ Upload at least one L2 machine report to start.")
